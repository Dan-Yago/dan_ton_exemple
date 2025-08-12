import pandas as pd
import matplotlib.pyplot as plt
from tabulate import tabulate
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

data=[]
#загрузка данных
file_path = 'data/2425/players_data-2024_2025.csv'
data.append(pd.read_csv(file_path))
for i in reversed(range(20,24)):
    file_path = f"data/gk/{i}_{i+1}.csv"
    data.append(pd.read_csv(file_path))

names,kol,mas,onxg=[],[],[],[]
head=['Name','PSxG', 'PSxG+/-']

for d in data:
    df=d
    poisk=df[(df['PSxG+/-'] >= 5)]
    kol.append(len(poisk))
    if not poisk.empty:
        for index, row in poisk.iterrows():
            names.append(row['Player'])
    else:
        print("Игрок не найден.")

players = list(set(names))
for name in players:
    temp=[]
    for d in data:

        df=d
        poisk=df[(df['Player'] == name)]
        if not poisk.empty:
            for index, row in poisk.iterrows():
                minutes_played = round(row['90s']*90,0)
                psxg=row['PSxG']
                age=row['Age']
                pm=row['PSxG+/-']
                cross=row['Stp%']
                mas.append([name,minutes_played,pm,cross])
                temp.append(pm)      
        else:
            #print("Игрок не найден.")
            mas.append([name,'-','-','-'])
            temp.append('-')
    onxg.append([name,temp[0],temp[1],temp[2],temp[3],temp[4]])

### сравнение данных
analiz=[]
for m in onxg:
    for i in range(2,6):
        if m[i]!='-':
            if m[i]>=5: #следующий сезон >= 5 psxg+
                analiz.append([m[0],m[i],m[i-1]])
plus,pl5=[],[]


for m in analiz:
    if m[2]!='-':
        if m[2]>=0:
            plus.append(m)
        if m[2]>=5:
            pl5.append(m)

# вывод след. сезон после 5 psxg, вывод плюсовых и вывод повторения >=5 psxg
headers=['Name','Seson 5+ psXG', 'Next season']
print (tabulate(analiz, tablefmt='pipe', stralign='center',headers=headers))
unique_values = list(dict.fromkeys(row[0] for row in plus))
print ('\n plus table   кол-во фамилий: ',len(unique_values))
print (tabulate(plus, tablefmt='pipe', stralign='center',headers=headers))
print ('\n pl5 table   кол-во: ',len(pl5))
print (tabulate(pl5, tablefmt='pipe', stralign='center',headers=headers))

### EXCEL
headers=['Name', 'Min', 'psXG+/-', 'Cross Stop%']
wb = Workbook()
ws = wb.active
colors = ['FFD3D3D3', 'FFFFFFFF']  # Серый и белый
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#заголовки
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.border = thin_border
# Определяем порядок цветов для игроков
player_colors = {}
current_color_idx = 0
last_player = None
# Заполняем данные
for row_num, row_data in enumerate(mas, 2):  # Начинаем со 2-й строки
    player_name = row_data[0]
    
    # Определяем цвет для игрока
    if player_name != last_player:
        current_color_idx = (current_color_idx + 1) % 2
        last_player = player_name
    
    fill = PatternFill(start_color=colors[current_color_idx],
                      end_color=colors[current_color_idx],
                      fill_type='solid')
    
    # Записываем данные
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.fill = fill
        cell.border = thin_border
# Автоподбор ширины столбцов
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col)
    adjusted_width = max_length + 2
    ws.column_dimensions[col[0].column_letter].width = adjusted_width

"""
#сохранение файла
wb.save('players_stats_colored.xlsx')
print("Файл успешно сохранен как players_stats_colored.xlsx")
"""

#рисуем картинку
fig = plt.figure(figsize=(13,9))
ax = plt.axes()

players, old_values, new_values=[],[],[]
for i in analiz:
    if i[2]!='-':
        players.append(i[0])
        old_values.append(i[1])
        new_values.append(i[2])
men0=0
bol0=0
for i in range(len(new_values)):
    if new_values[i]<0:
        men0+=1
    else:
        bol0+=1

for i, (player, old, new) in enumerate(zip(players, old_values, new_values)):
    if new>old:
        cols='#00AA00'
    else:
        cols = 'black'
    if players[i]==players[i-1]:
        plt.scatter(old, player+' ', color='blue', s=100)
        plt.scatter(new, player+' ', color='red', s=100)
        plt.annotate("", xy=(new, player+' '), xytext=(old, player+' '),
                     arrowprops=dict(arrowstyle="->", lw=2, color=cols))     # Толщ обводки))
    else:
        plt.scatter(old, player, color='blue', s=100)
        plt.scatter(new, player, color='red', s=100)
        plt.annotate("", xy=(new, player), xytext=(old, player),
                     arrowprops=dict(arrowstyle="->", lw=2, color=cols))

plt.axvline(x=0, color='red', linestyle='--', linewidth=1, alpha=0.5)     # Пунк
plt.title("Следующий сезон после psxG>5")
plt.xlabel("psXG - goals")
plt.grid(axis='x')
plt.tight_layout()
plt.savefig('img/psxg_probl.png',dpi=300)
plt.show()
